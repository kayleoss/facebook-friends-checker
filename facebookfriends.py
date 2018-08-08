from selenium import webdriver
import time, datetime
from openpyxl import load_workbook

chrome_options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_argument("--headless")
chrome_options.add_experimental_option("prefs",prefs)
chrome_path=r"path_to_chromedriver"
driver = webdriver.Chrome(chrome_path, chrome_options=chrome_options)

excelbook = load_workbook(filename="./friends.xlsx")
friends = []
discarded_txt = ["friends", "Facebook", "Home", "Find Friends", "Friend Requests", "Messages", "Notifications", "Account Settings"]

def main():
    driver.get("https://facebook.com/login")
    driver.find_element_by_id("email").send_keys("your_facebook_email@gmail.com")
    driver.find_element_by_id("pass").send_keys("your_facebook_password")
    driver.find_element_by_id("loginbutton").click()
    driver.get("https://www.facebook.com/your_facebook_url/friends")

    for i in range(8):
        driver.execute_script("window.scrollBy(0, 2500)")
        time.sleep(1)

    friends_list = driver.find_elements_by_xpath("//a[@data-gt]")
    print("GETTING ALL FRIENDS....")

    for friend in friends_list:
        if any([word in friend.text for word in discarded_txt]):
            pass
        else:
            global friends
            friends.append(friend.text)
    print len(friends)
    driver.quit()

def write_to_workbook():
    print("WRITING TO WORKBOOK......")
    time_now = str(datetime.date.today())
    excelbook.create_sheet(time_now)
    for friend in friends:
        excelbook[time_now].append([friend])
    print("----------- TASK COMPLETED ----------- ")
    excelbook.save("friends.xlsx")



if __name__ == "__main__":
    main()
    write_to_workbook()
